from collections import defaultdict
from datetime import date, timedelta
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app import models
from app.services.helpers import PAIR_SIZE_AH, _get_week_start, days


def _safe_sheet_name(base: str) -> str:
    name = base.replace(":", "-")[:31]
    return name if name else "Sheet"


def _apply_diff_formatting(ws):
    fill_added = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
    fill_removed = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red
    fill_changed = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # yellow
    fill_dim = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")  # gray
    bold_font = Font(bold=True)
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    if not headers:
        return
    col_type = headers.get("type")
    col_plan_subj = headers.get("plan_subject")
    col_plan_teach = headers.get("plan_teacher")
    col_plan_room = headers.get("plan_room")
    col_act_subj = headers.get("actual_subject")
    col_act_teach = headers.get("actual_teacher")
    col_act_room = headers.get("actual_room")
    # Freeze and autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # Column widths
    for h, c in headers.items():
        width = 12
        if h in ("group_name", "plan_subject", "actual_subject"):
            width = 24
        if h in ("plan_teacher", "actual_teacher"):
            width = 22
        if h in ("plan_room", "actual_room"):
            width = 18
        ws.column_dimensions[get_column_letter(c)].width = width
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=col_type).value if col_type else None
        if not t:
            continue
        if t == "added":
            for c in (col_act_subj, col_act_teach, col_act_room):
                if c:
                    ws.cell(row=r, column=c).fill = fill_added
                    ws.cell(row=r, column=c).font = bold_font
            for c in (col_plan_subj, col_plan_teach, col_plan_room):
                if c:
                    ws.cell(row=r, column=c).fill = fill_dim
        elif t == "removed":
            for c in (col_plan_subj, col_plan_teach, col_plan_room):
                if c:
                    ws.cell(row=r, column=c).fill = fill_removed
                    ws.cell(row=r, column=c).font = bold_font
            for c in (col_act_subj, col_act_teach, col_act_room):
                if c:
                    ws.cell(row=r, column=c).fill = fill_dim
        elif t == "changed":
            def _val(c):
                return ws.cell(row=r, column=c).value if c else None
            if col_plan_subj and col_act_subj and _val(col_plan_subj) != _val(col_act_subj):
                ws.cell(row=r, column=col_plan_subj).fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
                ws.cell(row=r, column=col_act_subj).fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
            if col_plan_teach and col_act_teach and _val(col_plan_teach) != _val(col_act_teach):
                ws.cell(row=r, column=col_plan_teach).fill = fill_changed
                ws.cell(row=r, column=col_act_teach).fill = fill_changed
            if col_plan_room and col_act_room and _val(col_plan_room) != _val(col_act_room):
                ws.cell(row=r, column=col_plan_room).fill = fill_changed
                ws.cell(row=r, column=col_act_room).fill = fill_changed



def _collect_weekly_slots_in_range(
    db: Session,
    start_date: date,
    end_date: date,
    group_names: Optional[List[str]] = None,
) -> List[Dict]:
    items: List[Dict] = []
    # Resolve group filter
    group_ids: Optional[set[int]] = None
    if group_names:
        ids = []
        for name in group_names:
            g = db.query(models.Group).filter(models.Group.name == name).first()
            if g:
                ids.append(g.id)
        group_ids = set(ids) if ids else set()
    # Fetch all weekly distributions intersecting the range
    q = (
        db.query(models.WeeklyDistribution)
        .filter(models.WeeklyDistribution.week_end >= start_date)
        .filter(models.WeeklyDistribution.week_start <= end_date)
    )
    if group_ids is not None:
        if not group_ids:
            return []
        q = q.join(models.ScheduleItem).filter(models.ScheduleItem.group_id.in_(group_ids))
    dists = q.all()
    for d in dists:
        item = d.schedule_item
        weekly_hours = d.hours_even if d.is_even_week else d.hours_odd
        pairs = int(weekly_hours // PAIR_SIZE_AH) if weekly_hours else 0
        for slot in (d.daily_schedule or []):
            try:
                day_idx = days.index(slot["day"])
            except ValueError:
                continue
            slot_date = d.week_start + timedelta(days=day_idx)
            if slot_date < start_date or slot_date > end_date:
                continue
            items.append(
                {
                    "date": slot_date,
                    "day": slot["day"],
                    "start_time": slot["start_time"],
                    "end_time": slot["end_time"],
                    "group_name": item.group.name,
                    "subject_name": item.subject.name,
                    "teacher_name": item.teacher.name,
                    "room_name": item.room.name,
                    "week_start": d.week_start,
                    "week_end": d.week_end,
                    "is_even_week": bool(d.is_even_week),
                    "weekly_hours": weekly_hours,
                    "weekly_pairs": pairs,
                }
            )
    # Sort primarily by group for clearer grouped view
    items.sort(key=lambda x: (x["group_name"], x["date"], x["start_time"]))
    return items


def build_plan_excel(db: Session, start_date: date, end_date: date, group_name: Optional[str] = None, group_names: Optional[List[str]] = None) -> BytesIO:
    names = None
    if group_names:
        names = group_names
    elif group_name:
        names = [group_name]
    rows = _collect_weekly_slots_in_range(db, start_date, end_date, names)
    df = pd.DataFrame(rows)
    # Summary per group/subject
    summary_rows: List[Dict] = []
    if names:
        q = db.query(models.ScheduleItem).join(models.Group).filter(models.Group.name.in_(names))
    else:
        q = db.query(models.ScheduleItem)
    items = q.all()
    for it in items:
        dists = (
            db.query(models.WeeklyDistribution)
            .filter(models.WeeklyDistribution.schedule_item_id == it.id)
            .filter(models.WeeklyDistribution.week_end >= start_date)
            .filter(models.WeeklyDistribution.week_start <= end_date)
            .all()
        )
        ah_assigned = sum((d.hours_even if d.is_even_week else d.hours_odd) or 0 for d in dists)
        summary_rows.append(
            {
                "group_name": it.group.name,
                "subject_name": it.subject.name,
                "teacher_name": it.teacher.name,
                "room_name": it.room.name,
                "total_hours": it.total_hours,
                "weekly_hours": it.weekly_hours,
                "week_type": it.week_type,
                "assigned_hours": ah_assigned,
                "remaining_hours": max(0.0, it.total_hours - ah_assigned),
            }
        )
    df_sum = pd.DataFrame(summary_rows)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        (df if not df.empty else pd.DataFrame(columns=[
            "date","day","start_time","end_time","group_name","subject_name","teacher_name","room_name","week_start","week_end","is_even_week","weekly_hours","weekly_pairs"
        ])).to_excel(writer, index=False, sheet_name="Weekly Plan")
        (df_sum if not df_sum.empty else pd.DataFrame(columns=[
            "group_name","subject_name","teacher_name","room_name","total_hours","weekly_hours","week_type","assigned_hours","remaining_hours"
        ])).to_excel(writer, index=False, sheet_name="Hours Summary")
        # Basic formatting
        ws1 = writer.sheets.get("Weekly Plan")
        if ws1 is not None:
            ws1.freeze_panes = "A2"
            ws1.auto_filter.ref = ws1.dimensions
        ws2 = writer.sheets.get("Hours Summary")
        if ws2 is not None:
            ws2.freeze_panes = "A2"
            ws2.auto_filter.ref = ws2.dimensions
    buf.seek(0)
    return buf


def _collect_day_actual(db: Session, date_: date, group_ids: Optional[set[int]] = None) -> Dict[Tuple[int, str], Dict]:
    ds = db.query(models.DaySchedule).filter(models.DaySchedule.date == date_).first()
    actual: Dict[Tuple[int, str], Dict] = {}
    if not ds:
        return actual
    for e in ds.entries:
        if group_ids is not None and ((not group_ids) or (e.group_id not in group_ids)):
            continue
        g = db.query(models.Group).get(e.group_id)
        s = db.query(models.Subject).get(e.subject_id)
        r = db.query(models.Room).get(e.room_id)
        t = db.query(models.Teacher).get(e.teacher_id) if e.teacher_id else None
        actual[(e.group_id, e.start_time)] = {
            "group_name": g.name if g else str(e.group_id),
            "start_time": e.start_time,
            "end_time": e.end_time,
            "subject_name": s.name if s else str(e.subject_id),
            "teacher_name": t.name if t else None,
            "room_name": r.name if r else str(e.room_id),
            "status": e.status,
        }
    return actual


def _collect_day_plan_from_weekly(db: Session, date_: date, group_ids: Optional[set[int]] = None) -> Dict[Tuple[int, str], Dict]:
    week_start = _get_week_start(date_)
    dow = days[date_.weekday()]
    dists = db.query(models.WeeklyDistribution).filter(models.WeeklyDistribution.week_start == week_start).all()
    plan: Dict[Tuple[int, str], Dict] = {}
    for d in dists:
        it = d.schedule_item
        if group_ids is not None and ((not group_ids) or (it.group_id not in group_ids)):
            continue
        for slot in d.daily_schedule or []:
            if slot.get("day") != dow:
                continue
            plan[(it.group_id, slot["start_time"])] = {
                "group_name": it.group.name,
                "start_time": slot["start_time"],
                "end_time": slot["end_time"],
                "subject_name": it.subject.name,
                "teacher_name": it.teacher.name,
                "room_name": it.room.name,
            }
    return plan


def build_day_with_diff_excel(db: Session, date_: date, group_name: Optional[str] = None, group_names: Optional[List[str]] = None) -> BytesIO:
    group_ids: Optional[set[int]] = None
    if group_names:
        ids = []
        for name in group_names:
            g = db.query(models.Group).filter(models.Group.name == name).first()
            if g:
                ids.append(g.id)
        group_ids = set(ids)
    elif group_name:
        g = db.query(models.Group).filter(models.Group.name == group_name).first()
        if g:
            group_ids = {g.id}
    actual = _collect_day_actual(db, date_, group_ids)
    plan = _collect_day_plan_from_weekly(db, date_, group_ids)

    # Build dataframes
    df_actual = pd.DataFrame(list(actual.values()))
    df_plan = pd.DataFrame(list(plan.values()))
    # Ensure consistent group-first sorting for readability
    if not df_actual.empty:
        df_actual.sort_values(by=["group_name", "start_time"], inplace=True)
    if not df_plan.empty:
        df_plan.sort_values(by=["group_name", "start_time"], inplace=True)
    # Diff
    keys = set(actual.keys()) | set(plan.keys())
    diff_rows: List[Dict] = []
    counters = defaultdict(int)
    for k in sorted(keys, key=lambda x: (x[0], x[1])):
        a = actual.get(k)
        p = plan.get(k)
        group_name_v = a["group_name"] if a else (p["group_name"] if p else "")
        start_time = a["start_time"] if a else (p["start_time"] if p else "")
        if a and not p:
            dtype = "added"
        elif p and not a:
            dtype = "removed"
        else:
            # both present
            changed = any([
                a.get("subject_name") != p.get("subject_name"),
                a.get("teacher_name") != p.get("teacher_name"),
                a.get("room_name") != p.get("room_name"),
            ])
            dtype = "changed" if changed else "same"
        counters[dtype] += 1
        diff_rows.append(
            {
                "group_name": group_name_v,
                "start_time": start_time,
                "type": dtype,
                "plan_subject": (p.get("subject_name") if p else None),
                "plan_teacher": (p.get("teacher_name") if p else None),
                "plan_room": (p.get("room_name") if p else None),
                "actual_subject": (a.get("subject_name") if a else None),
                "actual_teacher": (a.get("teacher_name") if a else None),
                "actual_room": (a.get("room_name") if a else None),
            }
        )
    df_diff = pd.DataFrame(diff_rows)

    # Hours summary per group
    group_set = sorted({r[0] for r in actual.keys()} | {r[0] for r in plan.keys()})
    summary_rows: List[Dict] = []
    for gid in group_set:
        g = db.query(models.Group).get(gid)
        pa = sum(1 for (gg, _st) in actual.keys() if gg == gid)
        pp = sum(1 for (gg, _st) in plan.keys() if gg == gid)
        summary_rows.append(
            {
                "group_name": g.name if g else str(gid),
                "actual_pairs": pa,
                "plan_pairs": pp,
                "delta_pairs": pa - pp,
                "actual_hours_AH": pa * PAIR_SIZE_AH,
                "plan_hours_AH": pp * PAIR_SIZE_AH,
                "delta_hours_AH": (pa - pp) * PAIR_SIZE_AH,
            }
        )
    df_summary = pd.DataFrame(summary_rows)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Write sheets
        (df_actual if not df_actual.empty else pd.DataFrame(columns=[
            "group_name","start_time","end_time","subject_name","teacher_name","room_name","status"
        ])).to_excel(writer, index=False, sheet_name="Day Actual")
        (df_plan if not df_plan.empty else pd.DataFrame(columns=[
            "group_name","start_time","end_time","subject_name","teacher_name","room_name"
        ])).to_excel(writer, index=False, sheet_name="Day Plan")
        (df_diff if not df_diff.empty else pd.DataFrame(columns=[
            "group_name","start_time","type","plan_subject","plan_teacher","plan_room","actual_subject","actual_teacher","actual_room"
        ])).to_excel(writer, index=False, sheet_name="Diff")
        (df_summary if not df_summary.empty else pd.DataFrame(columns=[
            "group_name","actual_pairs","plan_pairs","delta_pairs","actual_hours_AH","plan_hours_AH","delta_hours_AH"
        ])).to_excel(writer, index=False, sheet_name="Hours Summary")

        # Apply formatting for Diff sheet
        ws = writer.sheets.get("Diff")
        if ws is not None:
            _apply_diff_formatting(ws)

        # Optional: minor formatting on summary
        ws2 = writer.sheets.get("Hours Summary")
        if ws2 is not None:
            ws2.freeze_panes = "A2"
            ws2.auto_filter.ref = ws2.dimensions
            for c in range(1, ws2.max_column + 1):
                ws2.column_dimensions[get_column_letter(c)].width = 18

    buf.seek(0)
    return buf


def _collect_actual_slots_in_range(
    db: Session,
    start_date: date,
    end_date: date,
    group_names: Optional[List[str]] = None,
) -> List[Dict]:
    # Resolve group IDs
    group_ids: Optional[set[int]] = None
    if group_names:
        ids = []
        for name in group_names:
            g = db.query(models.Group).filter(models.Group.name == name).first()
            if g:
                ids.append(g.id)
        group_ids = set(ids) if ids else set()
    q = (
        db.query(models.DaySchedule)
        .filter(models.DaySchedule.date >= start_date)
        .filter(models.DaySchedule.date <= end_date)
    )
    day_plans = q.all()
    rows: List[Dict] = []
    for ds in day_plans:
        for e in ds.entries:
            if group_ids is not None and ((not group_ids) or (e.group_id not in group_ids)):
                continue
            g = db.query(models.Group).get(e.group_id)
            s = db.query(models.Subject).get(e.subject_id)
            r = db.query(models.Room).get(e.room_id)
            t = db.query(models.Teacher).get(e.teacher_id) if e.teacher_id else None
            day_str = days[ds.date.weekday()] if 0 <= ds.date.weekday() < len(days) else str(ds.date.weekday())
            rows.append(
                {
                    "date": ds.date,
                    "day": day_str,
                    "start_time": e.start_time,
                    "end_time": e.end_time,
                    "group_name": g.name if g else str(e.group_id),
                    "subject_name": s.name if s else str(e.subject_id),
                    "teacher_name": t.name if t else None,
                    "room_name": r.name if r else str(e.room_id),
                    "status": e.status,
                }
            )
    rows.sort(key=lambda x: (x["group_name"], x["date"], x["start_time"]))
    return rows


def _compute_diff_for_range(plan_rows: List[Dict], actual_rows: List[Dict]) -> List[Dict]:
    # Keys by (date, group_name, start_time)
    plan_map: Dict[tuple, Dict] = {}
    actual_map: Dict[tuple, Dict] = {}
    for r in plan_rows:
        plan_map[(r["date"], r["group_name"], r["start_time"])] = r
    for r in actual_rows:
        actual_map[(r["date"], r["group_name"], r["start_time"])] = r
    keys = sorted(set(plan_map.keys()) | set(actual_map.keys()))
    diffs: List[Dict] = []
    for k in keys:
        p = plan_map.get(k)
        a = actual_map.get(k)
        date_, group_name, start_time = k
        if p and not a:
            dtype = "removed"
        elif a and not p:
            dtype = "added"
        else:
            changed = any([
                a.get("subject_name") != p.get("subject_name"),
                a.get("teacher_name") != p.get("teacher_name"),
                a.get("room_name") != p.get("room_name"),
            ])
            dtype = "changed" if changed else "same"
        diffs.append(
            {
                "date": date_,
                "day": p.get("day") if p else a.get("day"),
                "group_name": group_name,
                "start_time": start_time,
                "type": dtype,
                "plan_subject": p.get("subject_name") if p else None,
                "plan_teacher": p.get("teacher_name") if p else None,
                "plan_room": p.get("room_name") if p else None,
                "actual_subject": a.get("subject_name") if a else None,
                "actual_teacher": a.get("teacher_name") if a else None,
                "actual_room": a.get("room_name") if a else None,
                "actual_status": a.get("status") if a else None,
            }
        )
    diffs.sort(key=lambda x: (x["group_name"], x["date"], x["start_time"]))
    return diffs


def build_schedule_range_excel(
    db: Session,
    start_date: date,
    end_date: date,
    group_names: Optional[List[str]] = None,
    view: str = "all",  # plan|actual|diff|all
    split_by_group: bool = False,
) -> BytesIO:
    plan_rows = _collect_weekly_slots_in_range(db, start_date, end_date, group_names)
    actual_rows = _collect_actual_slots_in_range(db, start_date, end_date, group_names)
    diff_rows = _compute_diff_for_range(plan_rows, actual_rows)

    def _to_df(rows: List[Dict], columns: List[str]) -> pd.DataFrame:
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=columns)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if not split_by_group:
            if view in ("plan", "all"):
                df_plan = _to_df(plan_rows, [
                    "date","day","start_time","end_time","group_name","subject_name","teacher_name","room_name","week_start","week_end","is_even_week","weekly_hours","weekly_pairs"
                ])
                if not df_plan.empty:
                    df_plan.sort_values(by=["group_name", "date", "start_time"], inplace=True)
                df_plan.to_excel(writer, index=False, sheet_name="Plan")
                ws = writer.sheets.get("Plan")
                if ws is not None:
                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = ws.dimensions
            if view in ("actual", "all"):
                df_actual = _to_df(actual_rows, [
                    "date","day","start_time","end_time","group_name","subject_name","teacher_name","room_name","status"
                ])
                if not df_actual.empty:
                    df_actual.sort_values(by=["group_name", "date", "start_time"], inplace=True)
                df_actual.to_excel(writer, index=False, sheet_name="Actual")
                ws = writer.sheets.get("Actual")
                if ws is not None:
                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = ws.dimensions
            if view in ("diff", "all"):
                df_diff = _to_df(diff_rows, [
                    "date","day","group_name","start_time","type","plan_subject","plan_teacher","plan_room","actual_subject","actual_teacher","actual_room","actual_status"
                ])
                df_diff.to_excel(writer, index=False, sheet_name="Diff")
                ws = writer.sheets.get("Diff")
                if ws is not None:
                    _apply_diff_formatting(ws)
        else:
            groups = sorted({r["group_name"] for r in (plan_rows + actual_rows)})
            for gname in groups:
                g_plan = [r for r in plan_rows if r["group_name"] == gname]
                g_actual = [r for r in actual_rows if r["group_name"] == gname]
                g_diff = _compute_diff_for_range(g_plan, g_actual)
                if view in ("plan", "all"):
                    df_plan = _to_df(g_plan, [
                        "date","day","start_time","end_time","group_name","subject_name","teacher_name","room_name","week_start","week_end","is_even_week","weekly_hours","weekly_pairs"
                    ])
                    sheet = _safe_sheet_name(f"Plan - {gname}")
                    df_plan.to_excel(writer, index=False, sheet_name=sheet)
                    ws = writer.sheets.get(sheet)
                    if ws is not None:
                        ws.freeze_panes = "A2"
                        ws.auto_filter.ref = ws.dimensions
                if view in ("actual", "all"):
                    df_actual = _to_df(g_actual, [
                        "date","day","start_time","end_time","group_name","subject_name","teacher_name","room_name","status"
                    ])
                    sheet = _safe_sheet_name(f"Actual - {gname}")
                    df_actual.to_excel(writer, index=False, sheet_name=sheet)
                    ws = writer.sheets.get(sheet)
                    if ws is not None:
                        ws.freeze_panes = "A2"
                        ws.auto_filter.ref = ws.dimensions
                if view in ("diff", "all"):
                    df_diff = _to_df(g_diff, [
                        "date","day","group_name","start_time","type","plan_subject","plan_teacher","plan_room","actual_subject","actual_teacher","actual_room","actual_status"
                    ])
                    sheet = _safe_sheet_name(f"Diff - {gname}")
                    df_diff.to_excel(writer, index=False, sheet_name=sheet)
                    ws = writer.sheets.get(sheet)
                    if ws is not None:
                        _apply_diff_formatting(ws)
        # Hours Summary across range (group+subject)
        # Aggregate pairs/hours for plan vs actual
        def _agg(rows: List[Dict]) -> Dict[tuple[str, str], int]:
            b: Dict[tuple[str, str], int] = defaultdict(int)
            for r in rows:
                key = (r.get("group_name"), r.get("subject_name"))
                b[key] += 1
            return b
        plan_cnt = _agg(plan_rows)
        actual_cnt = _agg(actual_rows)
        hs_rows: List[Dict] = []
        keys = sorted(set(plan_cnt.keys()) | set(actual_cnt.keys()))
        for (gname, sname) in keys:
            pp = plan_cnt.get((gname, sname), 0)
            ap = actual_cnt.get((gname, sname), 0)
            hs_rows.append({
                "group_name": gname,
                "subject_name": sname,
                "plan_pairs": pp,
                "actual_pairs": ap,
                "delta_pairs": ap - pp,
                "plan_hours_AH": pp * PAIR_SIZE_AH,
                "actual_hours_AH": ap * PAIR_SIZE_AH,
                "delta_hours_AH": (ap - pp) * PAIR_SIZE_AH,
            })
        df_hours = pd.DataFrame(hs_rows) if hs_rows else pd.DataFrame(columns=[
            "group_name","subject_name","plan_pairs","actual_pairs","delta_pairs","plan_hours_AH","actual_hours_AH","delta_hours_AH"
        ])
        df_hours.to_excel(writer, index=False, sheet_name="Hours Summary")
        ws = writer.sheets.get("Hours Summary")
        if ws is not None:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
    buf.seek(0)
    return buf
